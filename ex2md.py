#!/usr/bin/env python3

""" Made by SpeerSec - free for use, credit if modified or reused """


import argparse
import openpyxl
import magic

# Parse the command-line arguments
parser = argparse.ArgumentParser()
parser.add_argument("-f", "--file", required=True, help="name of the Excel file")
parser.add_argument("-s", "--sheet", help="name of the sheet to convert (default is the first sheet) Name is case sensitive.")
parser.add_argument("-r", "--range", help="range of cells to include in the table (e.g. A1:G16)")
parser.add_argument("-H","--headers", action="store_true", help="include the first row as the column headers")
parser.add_argument("-q","--quiet", action="store_true", help="do not print the markdown table")
parser.add_argument("-o","--output", help="override the default file output name")
args, remaining = parser.parse_known_args()

# If the help argument is given or no arguments were given, print the help menu
if remaining:
    parser.print_help()
    exit()

# Check the file extension
if not args.file.endswith((".xlsx", ".xml")):
    print("Error: only Excel files with the .xlsx or .xml extension are allowed.")
    exit()


# Open the file in binary mode
with open(args.file, "rb") as t:
    file_type = magic.from_buffer(t.read())
    t.close

# Check if the file is an Excel file
if "Microsoft Excel" not in file_type and "XML" not in file_type:
    print("Error: the provided file is not an Excel file.")
    exit()

if args.output:
    fileName = args.output
else:
    fileName = (str(args.file)).split('.')[0] + '.md'

# Open the Excel workbook and evaluate cell formulas
workbook = openpyxl.load_workbook(args.file, data_only=True)

# Select the sheet you want to convert
if args.sheet:
    sheet = workbook[args.sheet]
else:
    sheet = workbook.active

# Initialize the Markdown table
markdown_table = ""

# Determine the range of rows and columns to include in the table
if args.range:
    start_column, start_row = openpyxl.utils.cell.coordinate_from_string(args.range.split(":")[0])
    end_column, end_row = openpyxl.utils.cell.coordinate_from_string(args.range.split(":")[1])
    start_column = openpyxl.utils.cell.column_index_from_string(start_column)
    end_column = openpyxl.utils.cell.column_index_from_string(end_column)
else:
    start_row = 1
    start_column = 1
    end_row = sheet.max_row
    end_column = sheet.max_column


# Add the table headers if specified
if args.headers:
    # Initialize the header string for the Markdown table
    header_string = ""
    num_headers = 0

    # Add the row to the Markdown table
    for column in range(start_column, end_column + 1):
        cell = sheet.cell(row=start_row, column=column)
        # Check if the cell contains no value and skip it if it does
        if cell.value is None:
            continue
        header_string += f"| {cell.value} "
        num_headers += 1

    # Add the header string to the Markdown table
    markdown_table += header_string + "|\n"

    # Add the table divider
    markdown_table += "| --- " * num_headers + " |\n"

    # Increment the start row to skip the first row
    start_row += 1

# Add the table rows
for row in range(start_row, end_row + 1):
    # Initialize the row string for the Markdown table
    row_string = ""

    # Check if the entire row contains 'None' values
    row_contains_none = all(cell.value is None for cell in [sheet.cell(row=row, column=column) for column in range(start_column, end_column + 1)])

    # Skip the row if it contains only 'None' values
    if row_contains_none:
        continue

    # Add the row to the Markdown table
    for column in range(start_column, end_column + 1):
        # Check if the entire column contains 'None' values
        column_contains_none = all(cell.value is None for cell in [sheet.cell(row=row_num, column=column) for row_num in range(start_row, end_row + 1)])

        # Skip the column if it contains only 'None' values
        if column_contains_none:
            continue

        # Add the cell value to the row string
        cell = sheet.cell(row=row, column=column)
        row_string += f"| {cell.value} "

    # Add the row string to the Markdown table
    markdown_table += row_string + "|\n"

    
# Print the Markdown table
if not args.quiet:
    print(markdown_table)

# Write the Markdown table to a new .md file
with open(fileName, "w") as f:
    f.write(markdown_table)
    f.close
